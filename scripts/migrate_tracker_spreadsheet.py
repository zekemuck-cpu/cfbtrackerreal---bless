#!/usr/bin/env python3
"""
Migrate an old CFB Tracker spreadsheet (.xlsx) to a Dynasty Tracker
JSON file that can be imported via the app's "Import File" button.

Usage:
    python3 migrate_tracker_spreadsheet.py <path-to-xlsx> [output.json]

Pulls EVERYTHING per-year from each YYYY sheet:
- Schedule + scores + ranks + site (regular / CC / bowls)
- Final Top 25 polls (Media + Coaches)
- CFP bracket (seeds + auto-bid + conference)
- Conference standings (full table per conference)
- Conference championships (champion/opponent/coach)
- National championship (champion/opponent/coach)
- Awards detail (player + position + team + class)
- All-Americans (First / Second / Freshman teams)
- All-Conference (First / Second / Freshman teams, user's conference)
- Player Departures (UDFA + transfer reason + transfer destination)
- Recruiting class (per year, with archetype + ranks + dev trait + prev team)
- Players-of-the-Week list (conference / national, per week)

Also auto-detects the user's TeamBuilder teams by comparing the user's
conference rosters against the static FBS conference map and pairs
each TB with a removed FBS team in the same conference, so each TB
takes over a real FBS slot on import.

Requires: openpyxl  (pip install openpyxl)
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl


# ---- static FBS team registry (parsed from teamRegistry.js) -----------

def load_static_teams():
    """Parse src/data/teamRegistry.js → { tid: {tid, abbr, name, primaryColor,
    secondaryColor, logo, isFCS} } so the migrator can write
    dynasty.teams[tid] directly with the TB info merged into the
    replaced FBS slot."""
    here = Path(__file__).resolve().parent.parent
    src = (here / 'src' / 'data' / 'teamRegistry.js').read_text()
    start = src.find('export const TEAMS = {') + len('export const TEAMS = {')
    end = src.find('\n}', start)
    section = src[start:end]
    blocks = re.split(r'\n\s*(\d+):\s*\{', section)
    teams = {}
    for i in range(1, len(blocks), 2):
        tid = int(blocks[i])
        body = blocks[i + 1]
        cut = body.find('\n  },')
        if cut == -1:
            cut = body.find('\n  }')
        body = body[:cut] if cut != -1 else body
        fields = {'tid': tid}
        for fm in re.finditer(
            r'^\s*(\w+):\s*("(?:[^"\\]|\\.)*"|true|false|\d+)',
            body, re.MULTILINE,
        ):
            k, raw = fm.group(1), fm.group(2)
            if raw == 'true':
                v = True
            elif raw == 'false':
                v = False
            elif raw.startswith('"'):
                v = raw[1:-1]
            else:
                try:
                    v = int(raw)
                except ValueError:
                    v = raw
            fields[k] = v
        teams[tid] = fields
    return teams


STATIC_TEAMS = load_static_teams()
ABBR_TO_TID = {t['abbr']: t['tid'] for t in STATIC_TEAMS.values()}


# Class progression mirrors src/data/CLASS_PROGRESSION + the implicit
# Fr→So→Jr→Sr step. HS recruits enter as Fr; JUCO recruits start at
# their stated level. Sr / RS Sr graduate (no next year).
CLASS_PROGRESSION = {
    'HS': 'Fr',
    'JUCO Fr': 'Fr',
    'JUCO So': 'So',
    'JUCO Jr': 'Jr',
    'JUCO Sr': 'Sr',
    'Fr': 'Fr',
    'RS Fr': 'RS Fr',
    'So': 'So',
    'RS So': 'RS So',
    'Jr': 'Jr',
    'RS Jr': 'RS Jr',
    'Sr': 'Sr',
    'RS Sr': 'RS Sr',
}
NEXT_CLASS = {
    'Fr': 'So',
    'So': 'Jr',
    'Jr': 'Sr',
    'Sr': None,
    'RS Fr': 'RS So',
    'RS So': 'RS Jr',
    'RS Jr': 'RS Sr',
    'RS Sr': None,
}


# Spreadsheet uses short bowl names ("Rose", "Fiesta") while the app
# stores the full name ("Rose Bowl", "Fiesta Bowl"). The CFP-specific
# entries (CFP First Round, Nat'l Championship) need different
# game-type tagging in addition to a name swap.
BOWL_NAME_MAP = {
    "68 Ventures": "68 Ventures Bowl",
    "Alamo": "Alamo Bowl",
    "Arizona": "Arizona Bowl",
    "Armed Forces": "Armed Forces Bowl",
    "Bahamas": "Bahamas Bowl",
    "Birmingham": "Birmingham Bowl",
    "Boca Raton": "Boca Raton Bowl",
    "Citrus": "Citrus Bowl",
    "Cotton": "Cotton Bowl",
    "Cure": "Cure Bowl",
    "Duke's Mayo": "Duke's Mayo Bowl",
    "Famous Idaho Potato": "Famous Idaho Potato Bowl",
    "Fenway": "Fenway Bowl",
    "Fiesta": "Fiesta Bowl",
    "First Responder": "First Responder Bowl",
    "Frisco": "Frisco Bowl",
    "GameAbove Sports": "GameAbove Sports Bowl",
    "Gasparilla": "Gasparilla Bowl",
    "Gator": "Gator Bowl",
    "Hawaii": "Hawaii Bowl",
    "Holiday": "Holiday Bowl",
    "Independence": "Independence Bowl",
    "LA": "LA Bowl",
    "Las Vegas": "Las Vegas Bowl",
    "Liberty": "Liberty Bowl",
    "Military": "Military Bowl",
    "Music City": "Music City Bowl",
    "Myrtle Beach": "Myrtle Beach Bowl",
    "New Mexico": "New Mexico Bowl",
    "New Orleans": "New Orleans Bowl",
    "Orange": "Orange Bowl",
    "Peach": "Peach Bowl",
    "Pop-Tarts": "Pop-Tarts Bowl",
    "Rate": "Rate Bowl",
    "Reliaquest": "Reliaquest Bowl",
    "Rose": "Rose Bowl",
    "Salute to Veterans": "Salute to Veterans Bowl",
    "Sugar": "Sugar Bowl",
    "Sun": "Sun Bowl",
    "Texas": "Texas Bowl",
    # Special — CFP final keeps no "Bowl" suffix, matching app shape.
    "Nat'l Championship": "National Championship",
}


# CFP slot mapping. The bowl name (after normalization) tells us which
# CFP round + slot it represents — but only if the user is in the CFP
# bracket for that year. The app constructs CFP game IDs as
# `{slotId}-{year}` and links to them from the CFP Bracket and
# TeamYear pages, so we have to match that format.
BOWL_TO_CFP_SLOT = {
    'Sugar Bowl': ('cfpqf1', 'cfp_quarterfinal'),
    'Orange Bowl': ('cfpqf2', 'cfp_quarterfinal'),
    'Rose Bowl': ('cfpqf3', 'cfp_quarterfinal'),
    'Cotton Bowl': ('cfpqf4', 'cfp_quarterfinal'),
    'Peach Bowl': ('cfpsf1', 'cfp_semifinal'),
    'Fiesta Bowl': ('cfpsf2', 'cfp_semifinal'),
    'National Championship': ('cfpnc', 'cfp_championship'),
}


def normalize_bowl_name(short):
    """Map a Tracker short bowl name to the app's canonical full name.
    If the name is already a full name (already ends with 'Bowl' or
    matches a known entry), return as-is. Returns None for empty input."""
    if not short:
        return None
    s_ = short.strip()
    if not s_:
        return None
    if s_ in BOWL_NAME_MAP:
        return BOWL_NAME_MAP[s_]
    # Already-canonical (user typed 'Rose Bowl' explicitly, etc.)
    if s_.endswith(' Bowl') or s_ == 'National Championship':
        return s_
    # Best-effort fallback — append " Bowl" so it at least looks right
    return f'{s_} Bowl'


# ---- helpers ----------------------------------------------------------

def num(v):
    """Try to coerce v to int. Return None if not numeric."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def fnum(v):
    """Try to coerce v to float. Return None if not numeric."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def s(v):
    """Stringify v cleanly."""
    if v is None:
        return ''
    return str(v).strip()


def ranknum(v):
    """Parse '#24' or 24 into 24. Return None if not parseable."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    m = re.match(r'^#?(\d+)$', str(v).strip())
    return int(m.group(1)) if m else None


def stars_count(cell):
    """Count star symbols in a string. ☆ ★ * etc."""
    if not cell:
        return None
    raw = str(cell)
    n = 0
    for ch in raw:
        if ch in '☆★⭐':
            n += 1
    return n if n else None


def cell(row, idx):
    """Safe row indexing → empty string if past end."""
    if idx is None or idx >= len(row):
        return ''
    v = row[idx]
    return v


# Static FBS conference map (mirrors src/data/conferenceTeams.js so the
# script stays runnable without importing the JS source).
STATIC_CONFERENCES = {
    'ACC': ['BC', 'CAL', 'CLEM', 'DUKE', 'FSU', 'GT', 'LOU', 'MIA', 'NCST', 'UNC', 'PITT', 'SMU', 'SYR', 'STAN', 'UVA', 'VT', 'WAKE'],
    'Big Ten': ['ILL', 'IU', 'IOWA', 'UMD', 'MICH', 'MSU', 'MINN', 'NEB', 'NU', 'OSU', 'ORE', 'PSU', 'PUR', 'RUTG', 'UCLA', 'USC', 'WASH', 'WIS'],
    'Big 12': ['ARIZ', 'ASU', 'BU', 'BYU', 'UC', 'COLO', 'UH', 'ISU', 'KU', 'KSU', 'OKST', 'TCU', 'TTU', 'UCF', 'UTAH', 'WVU'],
    'SEC': ['BAMA', 'ARK', 'AUB', 'FLA', 'UGA', 'UK', 'LSU', 'MISS', 'MSST', 'MIZ', 'OU', 'SCAR', 'UT', 'TEX', 'TAMU', 'VAN'],
    'Pac-12': ['ORST', 'WSU'],
    'American': ['ARMY', 'CHAR', 'ECU', 'FAU', 'MEM', 'NAVY', 'UNT', 'RICE', 'TEM', 'TULN', 'TLSA', 'UAB', 'USF', 'UTSA'],
    'Mountain West': ['AFA', 'BOIS', 'CSU', 'FRES', 'HAW', 'NEV', 'SDSU', 'SJSU', 'UNM', 'UNLV', 'USU', 'WYO'],
    'Sun Belt': ['APP', 'ARST', 'CCU', 'GASO', 'GSU', 'JMU', 'JKST', 'ULM', 'UL', 'MRSH', 'ODU', 'USA', 'USM', 'TXST', 'TROY'],
    'MAC': ['AKR', 'BALL', 'BGSU', 'BUFF', 'CMU', 'EMU', 'KENT', 'M-OH', 'NIU', 'OHIO', 'TOL', 'WMU'],
    'Conference USA': ['DEL', 'FIU', 'KENN', 'LIB', 'LT', 'MTSU', 'MZST', 'NMSU', 'SHSU', 'UTEP', 'WKU'],
    'Independent': ['ND', 'CONN', 'MASS'],
}

ALL_FBS = {abbr for confs in STATIC_CONFERENCES.values() for abbr in confs}

# Tracker → app conference label normalization
CONF_LABEL_MAP = {
    'ACC': 'ACC',
    'American': 'American',
    'Big 12': 'Big 12',
    'Big Ten': 'Big Ten',
    'CUSA': 'Conference USA',
    'C-USA': 'Conference USA',
    'Conference USA': 'Conference USA',
    'MAC': 'MAC',
    'MWC': 'Mountain West',
    'Mountain West': 'Mountain West',
    'Pac-12': 'Pac-12',
    'PAC-12': 'Pac-12',
    'SEC': 'SEC',
    'Sun Belt': 'Sun Belt',
    'Independents': 'Independent',
    'Independent': 'Independent',
    'Independant': 'Independent',  # tolerate sheet typo
}


def normalize_conf(name):
    if not name:
        return None
    return CONF_LABEL_MAP.get(s(name), s(name))


# ---- year-sheet parsers -----------------------------------------------

def parse_schedule(rows, user_abbr, year, next_id_ref, make_game):
    """Parse left-side schedule rows (week | _ | bowlname | UserTeam | UserRank | CPUTeam | CPURank | Site | UserScore | CPUScore | Result)."""
    games = []
    for r in rows:
        if len(r) < 11:
            continue
        week_cell = r[1]
        week_label = s(week_cell)
        user_score = num(r[8])
        cpu_score = num(r[9])
        if user_score is None and cpu_score is None:
            continue
        opp = s(r[5])
        if not opp:
            continue
        user_rank = ranknum(r[4])
        opp_rank = ranknum(r[6])
        site = s(r[7]) or None
        bowl_name = s(r[2]) or None

        wnum = num(week_cell)
        is_cc = 'conf' in week_label.lower() and 'champ' in week_label.lower()
        is_bowl = 'bowl' in week_label.lower()

        if wnum is not None:
            games.append(make_game(
                wnum, opp, user_score, cpu_score, year,
                game_type='regular',
                user_rank=user_rank, opp_rank=opp_rank, site=site,
            ))
        elif is_cc:
            games.append(make_game(
                14, opp, user_score, cpu_score, year,
                game_type='conference_championship',
                is_cc=True,
                user_rank=user_rank, opp_rank=opp_rank, site=site,
            ))
        elif is_bowl:
            m = re.search(r'wk\s*(\d+)', week_label.lower())
            bowl_week = 14 + (int(m.group(1)) if m else 1)
            # Special CFP cases — these aren't real "bowls" in the app's
            # game-type taxonomy, they're CFP rounds. The bowl-name cell
            # tells us which round.
            raw_bowl = bowl_name or ''
            if raw_bowl.strip().lower() in ('cfp first round', 'cfp first', 'first round'):
                games.append(make_game(
                    bowl_week, opp, user_score, cpu_score, year,
                    game_type='cfp_first_round',
                    is_bowl=False, bowl_name=None, is_playoff=True,
                    user_rank=user_rank, opp_rank=opp_rank, site=site,
                ))
            elif raw_bowl.strip().lower() in ("nat'l championship", "national championship"):
                games.append(make_game(
                    bowl_week, opp, user_score, cpu_score, year,
                    game_type='cfp_championship',
                    is_bowl=True,  # Bowl page also surfaces championship
                    bowl_name=normalize_bowl_name(raw_bowl),
                    is_playoff=True,
                    user_rank=user_rank, opp_rank=opp_rank, site=site,
                ))
            else:
                games.append(make_game(
                    bowl_week, opp, user_score, cpu_score, year,
                    game_type='bowl',
                    is_bowl=True,
                    bowl_name=normalize_bowl_name(raw_bowl),
                    user_rank=user_rank, opp_rank=opp_rank, site=site,
                ))
    return games


def find_section(rows, label, col=1, max_row=None):
    """Find row index where rows[i][col] == label (case-insensitive)."""
    needle = label.lower()
    upper = max_row if max_row is not None else len(rows)
    for i in range(upper):
        r = rows[i]
        if len(r) > col:
            if isinstance(r[col], str) and r[col].strip().lower() == needle:
                return i
    return None


def parse_final_polls(rows):
    """Cols 1-4: Media (rank, team), Coaches (rank, team)."""
    start = find_section(rows, 'Final Top 25 Polls')
    if start is None:
        return None
    media = []
    coaches = []
    # Header at start+1, data starts start+2
    for i in range(start + 2, min(start + 30, len(rows))):
        r = rows[i]
        if len(r) < 5:
            continue
        m_rank = num(r[1])
        m_team = s(r[2])
        c_rank = num(r[3])
        c_team = s(r[4])
        if m_rank and m_team:
            media.append({'rank': m_rank, 'team': m_team.upper(), 'tid': None})
        if c_rank and c_team:
            coaches.append({'rank': c_rank, 'team': c_team.upper(), 'tid': None})
        if not (m_rank or c_rank):
            # Stop on first fully empty row in this band
            if i > start + 5:
                break
    if not media and not coaches:
        return None
    return {'media': media, 'coaches': coaches}


def parse_cfp_bracket(rows):
    """Cols 6-9: seed, team, conference, auto-bid (in 'College Football Playoff Bracket' section)."""
    start = find_section(rows, 'College Football Playoff Bracket', col=6)
    if start is None:
        return None
    seeds = []
    for i in range(start + 2, min(start + 18, len(rows))):
        r = rows[i]
        if len(r) < 10:
            continue
        seed = num(r[6])
        team = s(r[7])
        conf = s(r[8])
        auto = r[9]
        if seed and team:
            seeds.append({
                'seed': seed,
                'team': team.upper(),
                'conference': normalize_conf(conf),
                'autoBid': bool(auto) if auto is not None else False,
                'tid': None,
            })
    return seeds or None


def parse_conference_standings(rows):
    """Walk the multiple 'Conference / Conf. Rank / Team / Wins / Losses ...' tables."""
    standings = {}
    i = 0
    while i < len(rows):
        r = rows[i]
        if len(r) > 9 and isinstance(r[1], str) and r[1].strip() == 'Conference' and isinstance(r[2], str) and r[2].strip() == 'Conf. Rank':
            # Data follows on i+1 ...
            j = i + 1
            while j < len(rows):
                rr = rows[j]
                if len(rr) < 10:
                    j += 1
                    continue
                conf_label = s(rr[1])
                team = s(rr[3])
                if not conf_label:
                    break  # blank row separates tables
                if not team:
                    j += 1
                    continue
                conf_norm = normalize_conf(conf_label)
                wins = num(rr[4]) or 0
                losses = num(rr[5]) or 0
                pf = num(rr[7])
                pa = num(rr[8])
                # Point Diff stored as string like '+255' / '-50' / 0; coerce.
                diff_raw = rr[9]
                diff = None
                if isinstance(diff_raw, (int, float)):
                    diff = int(diff_raw)
                elif isinstance(diff_raw, str):
                    m = re.match(r'^([+-]?\d+)$', diff_raw.strip())
                    if m:
                        diff = int(m.group(1))
                conf_list = standings.setdefault(conf_norm, [])
                conf_list.append({
                    'team': team.upper(),
                    'wins': wins,
                    'losses': losses,
                    'pointsFor': pf,
                    'pointsAgainst': pa,
                    'pointDiff': diff,
                })
                j += 1
            i = j
            continue
        i += 1
    return standings or None


def parse_conference_championships(rows):
    """Right side, 'Conference Championships' header (col 19), data Conference, Champion, Opponent, Winning Coach."""
    start = find_section(rows, 'Conference Championships', col=19)
    if start is None:
        return None
    out = []
    for i in range(start + 2, min(start + 20, len(rows))):
        r = rows[i]
        if len(r) < 23:
            continue
        conf = s(r[19])
        champ = s(r[20])
        opp = s(r[21])
        coach = s(r[22])
        if not conf:
            break
        if champ:
            out.append({
                'conference': normalize_conf(conf),
                'champion': champ.upper(),
                'opponent': opp.upper() if opp else None,
                'coach': coach or None,
            })
    return out or None


def parse_national_championship(rows):
    """Right side, 'National Championship' header (col 19), single row data
    at start+2 (start+1 is the column-name sub-header row)."""
    start = find_section(rows, 'National Championship', col=19)
    if start is None:
        return None
    data_idx = start + 2
    if data_idx >= len(rows):
        return None
    r = rows[data_idx]
    if len(r) < 23:
        return None
    champ = s(r[19])
    opp = s(r[20])
    coach = s(r[21])
    if not champ:
        return None
    if champ.lower() in ('champion', 'conference championships'):
        return None
    return {
        'champion': champ.upper(),
        'opponent': opp.upper() if opp else None,
        'coach': coach or None,
    }


def parse_awards(rows):
    """Right side, 'Awards' header (col 19): Award, Player, Position, Team, Class."""
    start = find_section(rows, 'Awards', col=19)
    if start is None:
        return None
    # These keys MUST match Awards.jsx's AWARD_ORDER list — typo
    # mismatches (e.g. daveyOBrien vs daveyObrien) make the award
    # silently disappear from the page.
    AWARD_KEY_MAP = {
        'Heisman': 'heisman',
        'Maxwell': 'maxwell',
        'Walter Camp': 'walterCamp',
        'Bear Bryant Coach of the Year': 'bearBryantCoachOfTheYear',
        "Davey O'Brien": 'daveyObrien',
        'Davey OBrien': 'daveyObrien',
        'Chuck Bednarik': 'chuckBednarik',
        'Bronco Nagurski': 'broncoNagurski',
        'Jim Thorpe': 'jimThorpe',
        'Doak Walker': 'doakWalker',
        'Fred Biletnikoff': 'fredBiletnikoff',
        'Biletnikoff': 'fredBiletnikoff',
        'Lombardi': 'lombardi',
        'Unitas Golden Arm': 'unitasGoldenArm',
        'Edge Rusher of the Year': 'edgeRusherOfTheYear',
        'Outland': 'outland',
        'John Mackey': 'johnMackey',
        'Broyles': 'broyles',
        'Dick Butkus': 'dickButkus',
        'Butkus': 'dickButkus',
        'Rimington': 'rimington',
        'Lou Groza': 'louGroza',
        'Ray Guy': 'rayGuy',
        'Returner of the Year': 'returnerOfTheYear',
    }
    out = {}
    # Header at start+1, data starts start+2
    for i in range(start + 2, min(start + 40, len(rows))):
        r = rows[i]
        if len(r) < 24:
            continue
        award = s(r[19])
        player = s(r[20])
        pos = s(r[21])
        team = s(r[22])
        klass = s(r[23])
        if not award:
            # If we hit blank then a later (non-awards) section, bail
            continue
        # Stop when we run into the next section header
        if award in ('College Football History Book', 'National Championship', 'Conference Championships'):
            break
        key = AWARD_KEY_MAP.get(award, None)
        if not key:
            # Camel-case fallback for unknown award names
            key = re.sub(r'[^A-Za-z0-9 ]+', '', award).strip().replace(' ', '')
            if not key:
                continue
            key = key[0].lower() + key[1:]
        if player:
            out[key] = {
                'player': player,
                'position': pos or None,
                'team': team.upper() if team else None,
                'class': klass or None,
            }
    return out or None


def parse_all_americans(rows):
    """Right side, 'All-Americans' (col 25): three sub-tables First / Second / Freshman.

    Header at start, sub-header at start+1 with three (Position/Player/Team/Class)
    blocks. Data spans start+2 to ~start+25.
    """
    start = find_section(rows, 'All-Americans', col=25)
    if start is None:
        return None
    blocks = [
        ('first', 25, 26, 27, 28),   # Position, Player, Team, Class
        ('second', 29, 30, 31, 32),
        ('freshman', 33, 34, 35, 36),
    ]
    # Data begins 3 rows after the section header (designation labels +
    # column-name sub-header are above it).
    out = []
    for designation, pos_c, name_c, team_c, class_c in blocks:
        for i in range(start + 3, min(start + 30, len(rows))):
            r = rows[i]
            if len(r) <= class_c:
                continue
            pos = s(r[pos_c])
            name = s(r[name_c])
            team = s(r[team_c])
            klass = s(r[class_c])
            # Skip column-name header rows literally
            if name.lower() == 'player' and pos.lower() == 'position':
                continue
            if not name:
                continue
            out.append({
                'designation': designation,
                'position': pos or None,
                'player': name,
                'school': team.upper() if team else None,
                'class': klass or None,
            })
    return out or None


def parse_all_conference(rows):
    """Right side, 'All-Conference' (col 25): same shape as All-Americans, lower
    in the sheet (around row 33+)."""
    start = find_section(rows, 'All-Conference', col=25)
    if start is None:
        return None
    blocks = [
        ('first', 25, 26, 27, 28),
        ('second', 29, 30, 31, 32),
        ('freshman', 33, 34, 35, 36),
    ]
    out = []
    for designation, pos_c, name_c, team_c, class_c in blocks:
        for i in range(start + 3, min(start + 35, len(rows))):
            r = rows[i]
            if len(r) <= class_c:
                continue
            pos = s(r[pos_c])
            name = s(r[name_c])
            team = s(r[team_c])
            klass = s(r[class_c])
            if name.lower() == 'player' and pos.lower() == 'position':
                continue
            if not name:
                continue
            out.append({
                'designation': designation,
                'position': pos or None,
                'player': name,
                'school': team.upper() if team else None,
                'class': klass or None,
            })
    return out or None


def parse_player_departures(rows):
    """Right side, 'Player Departures' (col 14):
    Headers at +2: Player | Position | Class | Overall | _ | Draft Round | Transfer Reason | Transfer Destination"""
    start = find_section(rows, 'Player Departures', col=14)
    if start is None:
        return None
    out = []
    for i in range(start + 2, min(start + 60, len(rows))):
        r = rows[i]
        if len(r) < 22:
            continue
        name = s(r[14])
        pos = s(r[15])
        klass = s(r[16])
        ovr = num(r[17])
        draft_round = s(r[19])  # 'UDFA' or '1', '2'... or empty
        transfer_reason = s(r[20])
        transfer_dest = s(r[21])
        if not name:
            # Blank row → end of table (could be padding)
            if i > start + 4:
                break
            continue
        # Skip header row if it slips through
        if name.lower() == 'player':
            continue

        # Determine reason/movement
        if draft_round:
            # Drafted (incl. UDFA) — tracker uses Draft Round column for both
            reason = 'Pro Draft'
            mov_type = 'declared_for_draft'
            destination = None
        elif transfer_reason:
            reason = transfer_reason  # Keep raw reason from Tracker
            mov_type = 'transferred_out'
            destination = transfer_dest.upper() if transfer_dest else None
        else:
            # Likely graduating senior with no draft tag
            reason = 'Graduating'
            mov_type = 'graduated'
            destination = None

        out.append({
            'name': name,
            'position': pos or None,
            'class': klass or None,
            'overall': ovr,
            'draftRound': draft_round or None,
            'reason': reason,
            'destination': destination,
            'movementType': mov_type,
        })
    return out or None


def parse_recruits(rows):
    """Right side, 'Recruiting' (col 23): full recruiting class table.

    Header at start+2: Player | Class | Position | Archetype | Stars |
    National Rank | State Rank | Position Rank | Height | Weight |
    Hometown | State | Gem/Bust | Dev Trait | Previous Team
    """
    start = find_section(rows, 'Recruiting', col=23)
    if start is None:
        return None
    out = []
    base = 23
    for i in range(start + 2, min(start + 80, len(rows))):
        r = rows[i]
        if len(r) <= base + 14:
            continue
        name = s(r[base + 0])
        if not name or name.lower() == 'player':
            if i > start + 4 and not name:
                break
            continue
        klass = s(r[base + 1]) or 'HS'
        pos = s(r[base + 2])
        archetype = s(r[base + 3])
        stars = stars_count(r[base + 4])
        nat = num(r[base + 5])
        st = num(r[base + 6])
        prk = num(r[base + 7])
        height = s(r[base + 8])
        weight = num(r[base + 9])
        hometown = s(r[base + 10])
        state_str = s(r[base + 11])
        gem = s(r[base + 12])
        dev = s(r[base + 13]) or 'Normal'
        prev = s(r[base + 14])

        non_portal = klass in ('HS', 'JUCO Fr', 'JUCO So', 'JUCO Jr')
        out.append({
            'name': name,
            'class': klass,
            'position': pos or None,
            'archetype': archetype or None,
            'stars': stars,
            'nationalRank': nat,
            'stateRank': st,
            'positionRank': prk,
            'height': height or None,
            'weight': weight,
            'hometown': hometown or None,
            'state': state_str or None,
            'gemBust': gem or None,
            'devTrait': dev or 'Normal',
            'previousTeam': prev or None,
            'isPortal': not non_portal,
        })
    return out or None


# Column → internal-stat-key mapping for the YYYY sheet's
# Individual Statistics table. Keys mirror the app's internal format
# (`statsByYear[year][category][field]` — see BOXSCORE_TO_INTERNAL_MAP
# in DynastyContext.jsx for canonical names).
STAT_COL_MAP = {
    'passing':    {49: 'cmp', 50: 'att', 51: 'yds', 52: 'td', 54: 'int', 62: 'lng', 63: 'sacks'},
    'rushing':    {65: 'car', 66: 'yds', 68: 'td', 70: 'twentyPlus', 71: 'bt',
                   72: 'yac', 73: 'lng', 74: 'fum'},
    'receiving':  {76: 'rec', 77: 'yds', 79: 'td', 81: 'lng', 82: 'rac', 84: 'drops'},
    'blocking':   {85: 'sacksAllowed'},
    'defense':    {86: 'soloTkl', 87: 'astTkl', 89: 'tfl', 90: 'sacks',
                   91: 'int', 92: 'intYds', 95: 'td', 96: 'pd',
                   98: 'ff', 99: 'fr'},
    'kicking':    {103: 'fgm', 104: 'fga', 106: 'lng', 107: 'xpm', 108: 'xpa',
                   110: 'fgm29', 111: 'fga29', 112: 'fgm39', 113: 'fga39',
                   114: 'fgm49', 115: 'fga49', 116: 'fgm50', 117: 'fga50',
                   118: 'kickoffs', 119: 'touchbacks',
                   121: 'fgb', 122: 'xpb'},
    'punting':    {123: 'punts', 124: 'yds', 126: 'netYds', 128: 'in20',
                   129: 'tb', 130: 'lng', 131: 'block'},
    'kickReturn': {132: 'ret', 133: 'yds', 135: 'td', 136: 'lng'},
    'puntReturn': {137: 'ret', 138: 'yds', 140: 'td', 141: 'lng'},
}


def parse_stat_category(row, col_map):
    """Pull one category (passing / rushing / etc.) from a roster row,
    returning a dict shaped {fieldKey: numericValue}. Always returns the
    full set of keys (zero-filled for missing cells), matching the
    shape used elsewhere in the app — the 'remove zero stats' cleanup
    step in DynastyContext drops all-zero entries on save."""
    out = {}
    for col, key in col_map.items():
        v = fnum(row[col]) if col < len(row) else None
        out[key] = v if v is not None else 0
    return out


def parse_year_roster(rows):
    """Right side, 'Individual Statistics' (col 41): per-year roster of the
    user's team.

    Layout:
        R1 col 41  : 'Individual Statistics'
        R3 col 41+ : section group labels
        R4         : column headers (Player, Position, Class, Dev Trait,
                     Overall Rating, Games Played, Games Started,
                     Snaps Played, then per-section stat columns)
        R5+        : data
    """
    # Header marker
    if not (len(rows) > 1 and len(rows[1]) > 41 and
            isinstance(rows[1][41], str) and
            'individual statistics' in rows[1][41].lower()):
        found = False
        for i in range(min(5, len(rows))):
            r = rows[i]
            if len(r) > 41 and isinstance(r[41], str) and 'individual statistics' in r[41].lower():
                found = True
                break
        if not found:
            return []

    out = []
    empty_streak = 0
    for i in range(5, min(160, len(rows))):
        r = rows[i]
        if len(r) <= 48:
            empty_streak += 1
            if empty_streak >= 5 and i > 10:
                break
            continue
        name = s(r[41])
        pos = s(r[42])
        if not name or not pos:
            empty_streak += 1
            if empty_streak >= 5 and i > 10:
                break
            continue
        empty_streak = 0
        klass = s(r[43])
        dev_trait = s(r[44]) or 'Normal'
        ovr = num(r[45])
        games_played = num(r[46]) or 0
        games_started = num(r[47]) or 0
        snaps = num(r[48]) or 0

        stats = {
            'gamesPlayed': games_played,
            'gamesStarted': games_started,
            'snapsPlayed': snaps,
            'passing':    parse_stat_category(r, STAT_COL_MAP['passing']),
            'rushing':    parse_stat_category(r, STAT_COL_MAP['rushing']),
            'receiving':  parse_stat_category(r, STAT_COL_MAP['receiving']),
            'blocking':   parse_stat_category(r, STAT_COL_MAP['blocking']),
            'defense':    parse_stat_category(r, STAT_COL_MAP['defense']),
            'kicking':    parse_stat_category(r, STAT_COL_MAP['kicking']),
            'punting':    parse_stat_category(r, STAT_COL_MAP['punting']),
            'kickReturn': parse_stat_category(r, STAT_COL_MAP['kickReturn']),
            'puntReturn': parse_stat_category(r, STAT_COL_MAP['puntReturn']),
        }
        # Did the player put up ANY non-zero numbers? If not, drop the
        # stats payload entirely so we don't bloat output for years
        # the user hasn't filled in (the app's cleanup pass would do
        # this on save anyway).
        has_any = False
        for k, v in stats.items():
            if isinstance(v, dict):
                if any(x for x in v.values()):
                    has_any = True
                    break
            elif v:
                has_any = True
                break

        out.append({
            'name': name,
            'position': pos,
            'class': klass or None,
            'devTrait': dev_trait,
            'overall': ovr,
            'stats': stats if has_any else None,
        })
    return out


def parse_players_of_week(rows):
    """Right side, 'Players of the Week (User Only)' (col 14):
    Conference (Name, Week) and National (Name, Week)."""
    start = find_section(rows, 'Players of the Week (User Only)', col=14)
    if start is None:
        return None
    out = {'conference': [], 'national': []}
    # Headers around start+2; data start+3 onward
    for i in range(start + 3, min(start + 25, len(rows))):
        r = rows[i]
        if len(r) < 18:
            continue
        c_name = s(r[14])
        c_week = r[15]
        n_name = s(r[16])
        n_week = r[17]
        if c_name:
            out['conference'].append({'player': c_name, 'week': s(c_week)})
        if n_name:
            out['national'].append({'player': n_name, 'week': s(n_week)})
        if not (c_name or n_name):
            if i > start + 5:
                break
    if not out['conference'] and not out['national']:
        return None
    return out


# ---- TB detection -----------------------------------------------------

def detect_team_builders(all_year_data, user_team_abbr):
    """Detect TB teams by collecting every team appearing in any year's
    standings and subtracting the static FBS map.

    Returns a dict: { tb_abbr -> { conference, replacesAbbr } }
    """
    # Step 1: collect every team mentioned + which conference the user
    # had them in at their FIRST appearance. TBs replace FBS slots in
    # their original conference; later realignment moves are tracked
    # year-by-year via conferenceByTeamYear, separately.
    user_teams = set()
    user_team_to_conf = {}
    # Walk oldest year first so first-seen (original) conference wins.
    for year in sorted(all_year_data.keys()):
        standings = all_year_data[year].get('standings') or {}
        for conf, teams in standings.items():
            for entry in teams:
                abbr = entry['team']
                user_teams.add(abbr)
                user_team_to_conf.setdefault(abbr, conf)

    # TBs = teams the user has but aren't in static FBS
    tb_abbrs = sorted(user_teams - ALL_FBS)
    # User team abbr should always be a TB (it's the user's slot)
    if user_team_abbr and user_team_abbr not in tb_abbrs and user_team_abbr not in ALL_FBS:
        tb_abbrs.append(user_team_abbr)

    # Removed FBS teams = static FBS teams the user doesn't have at all
    removed_fbs = sorted(ALL_FBS - user_teams)

    # Pair each TB with a removed FBS team in the same conference (best
    # effort). If no exact-conference match exists, pick any unused
    # removed FBS team.
    tb_conf_map = {tb: user_team_to_conf.get(tb) for tb in tb_abbrs}
    # Map static FBS abbr → its conference
    fbs_to_conf = {abbr: conf for conf, abbrs in STATIC_CONFERENCES.items() for abbr in abbrs}

    tb_pairings = {}
    used = set()
    # Iterate user team first so it gets first pick at a same-conference slot.
    ordered_tbs = ([user_team_abbr] if user_team_abbr in tb_abbrs else []) + \
                  [t for t in tb_abbrs if t != user_team_abbr]

    # Pass 1: prefer same-conference pairings
    for tb in ordered_tbs:
        conf = tb_conf_map[tb]
        candidate = next(
            (rm for rm in removed_fbs if rm not in used and fbs_to_conf.get(rm) == conf),
            None,
        )
        if candidate:
            tb_pairings[tb] = {'conference': conf, 'replacesAbbr': candidate}
            used.add(candidate)
    # Pass 2: any remaining TB → any unused removed FBS team
    for tb in ordered_tbs:
        if tb in tb_pairings:
            continue
        candidate = next((rm for rm in removed_fbs if rm not in used), None)
        if candidate:
            tb_pairings[tb] = {'conference': tb_conf_map[tb], 'replacesAbbr': candidate}
            used.add(candidate)
        else:
            # No removed FBS to pair with — fall back to GASO placeholder
            tb_pairings[tb] = {'conference': tb_conf_map[tb], 'replacesAbbr': 'GASO'}

    return tb_pairings, removed_fbs


# ---- main migration ---------------------------------------------------

def migrate(xlsx_path: Path, output_path: Path):
    print(f'Reading {xlsx_path} …')
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # ----- User team + start year (Team sheet) ---------------------------
    team_sheet = wb['Team']
    team_rows = [r for r in team_sheet.iter_rows(values_only=True)]
    user_team_abbr = None
    start_year = None
    last_active_year = None
    team_records = {}

    for row in team_rows:
        year = num(row[1]) if len(row) > 1 else None
        role = s(row[2]) if len(row) > 2 else ''
        school = s(row[3]) if len(row) > 3 else ''
        conf = s(row[5]) if len(row) > 5 else ''
        wins = num(row[8]) if len(row) > 8 else None
        losses = num(row[9]) if len(row) > 9 else None
        if year and school and (wins is not None or losses is not None):
            if start_year is None:
                start_year = year
                user_team_abbr = school
            last_active_year = year
            team_records[year] = {
                'wins': wins or 0,
                'losses': losses or 0,
                'conference': normalize_conf(conf),
                'role': role or 'HC',
            }

    if not user_team_abbr:
        sys.exit('Could not find user team in Team sheet')
    print(f'  → User team: {user_team_abbr}')
    print(f'  → Years: {start_year}–{last_active_year}  ({len(team_records)} seasons)')

    # ----- School full name + coach role ---------------------------------
    coach_position = 'HC'
    user_team_full_name = None
    for y in range(start_year, last_active_year + 1):
        sn = str(y)
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True, max_row=6):
            for i, c in enumerate(row):
                if isinstance(c, str) and c.strip().lower().startswith('current job'):
                    nearby = [s(row[j]) for j in range(i + 1, min(i + 10, len(row)))]
                    role_label = next(
                        (x for x in nearby if x in ('Head Coach', 'Offensive Coordinator', 'Defensive Coordinator')),
                        '',
                    )
                    coach_position = {
                        'Head Coach': 'HC',
                        'Offensive Coordinator': 'OC',
                        'Defensive Coordinator': 'DC',
                    }.get(role_label, 'HC')
                    after_at = False
                    for c2 in nearby:
                        if c2 == 'at':
                            after_at = True
                            continue
                        if after_at and c2 and not c2.endswith(':'):
                            user_team_full_name = c2
                            break
                    break
            if user_team_full_name:
                break
        if user_team_full_name:
            break
    if not user_team_full_name:
        user_team_full_name = f'{user_team_abbr} Dynasty Team'
    coach_name = '[Your Name]'
    print(f'  → School: {user_team_full_name}')
    print(f'  → Coach role: {coach_position}')

    # ----- Per-year structured data --------------------------------------
    print('\nExtracting per-year data from each YYYY sheet …')
    games = []
    next_game_id_box = [1]

    def make_game(week, opp, user_score, cpu_score, year, *,
                  game_type='regular', is_bowl=False, is_cc=False, bowl_name=None,
                  is_playoff=False,
                  user_rank=None, opp_rank=None, site=None):
        gid = f'imported-game-{next_game_id_box[0]}'
        next_game_id_box[0] += 1
        won = (user_score is not None and cpu_score is not None and user_score > cpu_score)
        location = (site or '').lower() if site else None
        is_home = location == 'home'
        is_away = location == 'away'
        is_neutral = location == 'neutral' or is_bowl or is_cc or is_playoff
        return {
            'id': gid,
            'week': week,
            'year': year,
            'gameType': game_type,
            'userTeam': user_team_abbr,
            'opponent': opp,
            'team1': user_team_abbr,
            'team2': opp,
            'team1Score': user_score if user_score is not None else 0,
            'team2Score': cpu_score if cpu_score is not None else 0,
            'team1Rank': user_rank,
            'team2Rank': opp_rank,
            'teamScore': user_score if user_score is not None else 0,
            'opponentScore': cpu_score if cpu_score is not None else 0,
            'userRank': user_rank,
            'opponentRank': opp_rank,
            'result': 'win' if won else ('loss' if (user_score is not None and cpu_score is not None) else None),
            'winner': user_team_abbr if won else opp,
            'isPlayed': user_score is not None and cpu_score is not None and (user_score > 0 or cpu_score > 0),
            'isBowlGame': is_bowl,
            'bowlGameName': bowl_name,
            'isConferenceChampionship': is_cc,
            'isPlayoff': is_playoff,
            'location': location,
            'isHome': is_home,
            'isAway': is_away,
            'isNeutral': is_neutral,
        }

    all_year_data = {}  # year -> dict of parsed sections

    for year in sorted(team_records.keys()):
        sn = str(year)
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))

        # Schedule + games[]
        year_games = parse_schedule(rows, user_team_abbr, year, next_game_id_box, make_game)
        games.extend(year_games)

        # Structured sections
        year_data = {
            'games': year_games,
            'standings': parse_conference_standings(rows),
            'finalPolls': parse_final_polls(rows),
            'cfpSeeds': parse_cfp_bracket(rows),
            'confChamps': parse_conference_championships(rows),
            'natlChamp': parse_national_championship(rows),
            'awards': parse_awards(rows),
            'allAmericans': parse_all_americans(rows),
            'allConference': parse_all_conference(rows),
            'departures': parse_player_departures(rows),
            'recruits': parse_recruits(rows),
            'potw': parse_players_of_week(rows),
            'roster': parse_year_roster(rows),
        }
        all_year_data[year] = year_data
        print(f'  → {year}: '
              f'{len(year_games)} games, '
              f'{len(year_data["standings"] or {})} confs, '
              f'{len(year_data["awards"] or {})} awards, '
              f'{len(year_data["allAmericans"] or [])} AA, '
              f'{len(year_data["allConference"] or [])} AC, '
              f'{len(year_data["departures"] or [])} departures, '
              f'{len(year_data["recruits"] or [])} recruits, '
              f'{len(year_data["roster"])} roster')

    # ----- Auto-detect TBs -----------------------------------------------
    print('\nAuto-detecting TeamBuilder teams …')
    tb_pairings, removed_fbs = detect_team_builders(all_year_data, user_team_abbr)
    print(f'  → TBs detected: {len(tb_pairings)} → {sorted(tb_pairings.keys())}')
    print(f'  → Removed FBS slots: {removed_fbs}')
    for tb, info in sorted(tb_pairings.items()):
        print(f'    {tb} → replaces {info["replacesAbbr"]} ({info["conference"]})')

    # ----- Build customTeams map -----------------------------------------
    # Hard-coded metadata for the seven TBs the user actually imported
    # (Stony Brook, Albany, Montana, Montana State, Tarleton State,
    #  UC Davis, New Hampshire). Anything not in this map falls back to
    # generic placeholders.
    TB_METADATA = {
        'STONY': {
            'name': 'Stony Brook Seawolves',
            'logoUrl': 'https://i.imgur.com/aeRqldX.png',
            'primaryColor': '#990000',
            'secondaryColor': '#16223e',
        },
        'ALB': {
            'name': 'Albany Great Danes',
            'logoUrl': 'https://i.imgur.com/Cu6gDNB.png',
            'primaryColor': '#46166b',
            'secondaryColor': '#fdb913',
        },
        'MONT': {
            'name': 'Montana Grizzlies',
            'logoUrl': 'https://i.imgur.com/7y9yT1w.png',
            'primaryColor': '#9d2235',
            'secondaryColor': '#a2aaad',
        },
        'MTST': {
            'name': 'Montana State Bobcats',
            'logoUrl': 'https://i.imgur.com/RgiOy5C.png',
            'primaryColor': '#003875',
            'secondaryColor': '#b69146',
        },
        'TARL': {
            'name': 'Tarleton State Texans',
            'logoUrl': 'https://i.imgur.com/bruXgzS.png',
            'primaryColor': '#3f1452',
            'secondaryColor': '#ffffff',
        },
        'UCD': {
            'name': 'UC Davis Aggies',
            'logoUrl': 'https://i.imgur.com/VxGSYa8.png',
            'primaryColor': '#022851',
            'secondaryColor': '#ffbf00',
        },
        'HAMP': {
            'name': 'New Hampshire Wildcats',
            'logoUrl': 'https://i.imgur.com/LORTtDv.png',
            'primaryColor': '#003591',
            'secondaryColor': '#ffffff',
        },
    }

    # Build dynasty.teams[tid] entries. Start from the full static FBS
    # map so the All Teams page (and every other consumer) sees every
    # team. For TBs we then override the replaced FBS team's slot
    # in-place: the tid stays the same as the FBS team's tid but the
    # abbr/name/colors/logo become the TB's. The site never has to
    # know it's a "teambuilder" — it just reads dynasty.teams[tid]
    # and shows what's there.
    dynasty_teams = {}
    for tid, t in STATIC_TEAMS.items():
        dynasty_teams[tid] = {
            **t,
            'byYear': {},
        }
    abbr_to_tid = dict(ABBR_TO_TID)  # working copy with TB swaps applied
    user_team_tid = None
    for tb, info in tb_pairings.items():
        meta = TB_METADATA.get(tb, {})
        replaced = info['replacesAbbr']
        slot_tid = ABBR_TO_TID.get(replaced)
        if slot_tid is None:
            print(f'  ! Cannot place {tb}: replacement abbr {replaced!r} not in static FBS map')
            continue
        static = STATIC_TEAMS.get(slot_tid, {})
        # No `isCustom` flag — the app shouldn't need to know this slot
        # used to belong to a different team. It's just dynasty.teams[tid]
        # like any other team, with whatever name/abbr/colors/logo are
        # written here. byYear is filled in below for record/schedule.
        merged = {
            'tid': slot_tid,
            'abbr': tb,
            'name': meta.get('name', tb),
            'primaryColor': meta.get('primaryColor', static.get('primaryColor', '#444444')),
            'secondaryColor': meta.get('secondaryColor', static.get('secondaryColor', '#ffffff')),
            'logo': meta.get('logoUrl', ''),
            'isFCS': False,
            'byYear': {},
        }
        if tb == user_team_abbr:
            merged['userId'] = 'currentUser'
        dynasty_teams[slot_tid] = merged
        # Remap abbr → tid so the TB's abbr resolves to the slot's tid
        # (and the original FBS abbr no longer resolves to anything).
        abbr_to_tid.pop(replaced, None)
        abbr_to_tid[tb] = slot_tid
        if tb == user_team_abbr:
            user_team_tid = slot_tid

    if user_team_tid is None:
        # User team isn't a TB — must be a real FBS team. Bring in the
        # static slot so the dashboard can find a userId-tagged team.
        user_team_tid = abbr_to_tid.get(user_team_abbr)
        if user_team_tid and user_team_tid in STATIC_TEAMS:
            base = dict(STATIC_TEAMS[user_team_tid])
            base['userId'] = 'currentUser'
            base.setdefault('byYear', {})
            dynasty_teams[user_team_tid] = base

    # Drop FBS slots the user removed and DIDN'T pair with a TB. This
    # collapses the team count down to whatever the user actually has
    # in their dynasty — no zombie "Akron Zips" entries when the user
    # removed AKR's slot and STONY took it over (handled above), and
    # no zombie "Kennesaw State" / "ULM" when the user just removed
    # them outright with no TB replacement.
    paired_replacements = {info['replacesAbbr'] for info in tb_pairings.values()}
    truly_removed = [a for a in removed_fbs if a not in paired_replacements]
    for abbr in truly_removed:
        slot_tid = ABBR_TO_TID.get(abbr)
        if slot_tid is not None:
            dynasty_teams.pop(slot_tid, None)
            abbr_to_tid.pop(abbr, None)
    if truly_removed:
        print(f'  → Truly removed FBS (no TB replacement): {truly_removed}')

    def tid_for(abbr):
        """Resolve any abbr to its tid (TB or FBS). Returns None if
        nothing matches."""
        if not abbr:
            return None
        return abbr_to_tid.get(s(abbr))

    # ----- Players: merge per-year rosters from each YYYY sheet ---------
    # The Individual sheet only has career-level data (one row per
    # player ever-on-team) and only carries OVR progression as a string,
    # no per-year roster membership. The YYYY sheets each carry the
    # actual roster for that year, so we walk those and merge.

    def split_name(full):
        """Split 'First Last' into (first, last). Multi-word last names
        keep everything after the first word as last name."""
        full = full.strip()
        if not full:
            return ('', '')
        parts = full.split()
        if len(parts) == 1:
            return (parts[0], '')
        return (parts[0], ' '.join(parts[1:]))

    # Pull supplemental Individual-sheet metadata (stars, hometown,
    # year_started, dev trait, etc.) keyed by name so we can enrich
    # roster-derived player records.
    ind = wb['Individual']
    ind_meta = {}
    for row in ind.iter_rows(values_only=True):
        if not row or len(row) < 13:
            continue
        team = s(row[1])
        name = s(row[2])
        position = s(row[3])
        school = s(row[4])
        if not (team and name and position and school):
            continue
        if school != user_team_abbr:
            continue
        year_started = num(row[6])
        stars_cell = s(row[7])
        stars = stars_cell.count('☆') if stars_cell else None
        nat_rank = num(row[10])
        dev_trait = s(row[12]) or 'Normal'
        ind_meta[name] = {
            'yearStarted': year_started,
            'stars': stars,
            'nationalRank': nat_rank,
            'devTrait': dev_trait,
        }

    # Walk each year's roster and merge. Same name = same player.
    # NOTE: tracker has typos like 'Holoway' vs 'Holloway' across years.
    # We trust each year's roster string and only merge on EXACT name.
    players_by_name = {}
    next_pid = 1

    for year in sorted(all_year_data.keys()):
        roster = all_year_data[year].get('roster') or []
        for entry in roster:
            name = entry['name']
            if name not in players_by_name:
                first, last = split_name(name)
                meta = ind_meta.get(name, {})
                pid = next_pid
                next_pid += 1
                players_by_name[name] = {
                    'pid': pid,
                    'id': f'imported-player-{pid}',
                    'name': name,
                    'firstName': first,
                    'lastName': last,
                    'position': entry['position'],
                    'team': user_team_tid,            # numeric tid
                    'school': user_team_abbr,
                    'year': entry.get('class') or 'Sr',
                    'stars': meta.get('stars'),
                    'nationalRank': meta.get('nationalRank'),
                    'devTrait': entry.get('devTrait') or meta.get('devTrait') or 'Normal',
                    'devTraitByYear': {},
                    'overall': entry.get('overall'),
                    'overallByYear': {},
                    'overallProgression': [],
                    'teamsByYear': {},                # tid (numeric)
                    'classByYear': {},
                    'statsByYear': {},
                    'movementByYear': {},
                    'movements': [],
                    'yearStarted': meta.get('yearStarted') or year,
                    'jerseyNumber': '',
                    'archetype': '',
                    'height': '',
                    'weight': None,
                    'hometown': '',
                    'state': '',
                    'pictureUrl': '',
                    'isHonorOnly': False,
                    'isPortal': False,
                    'isRecruit': False,
                }
            p = players_by_name[name]
            ystr = str(year)
            # Numeric tid — the app's roster check expects this.
            p['teamsByYear'][ystr] = user_team_tid
            if entry.get('overall') is not None:
                p['overallByYear'][ystr] = entry['overall']
            if entry.get('class'):
                p['classByYear'][ystr] = entry['class']
            if entry.get('devTrait'):
                p['devTraitByYear'][ystr] = entry['devTrait']
            if entry.get('stats'):
                p['statsByYear'][ystr] = entry['stats']
            # Take the latest year's position / dev trait as canonical
            p['position'] = entry['position']
            if entry.get('devTrait'):
                p['devTrait'] = entry['devTrait']
            if entry.get('class'):
                p['year'] = entry['class']
            # Final overall = highest filled value
            if entry.get('overall') is not None and (
                    p['overall'] is None or entry['overall'] >= p['overall']):
                p['overall'] = entry['overall']

    # Build overallProgression from sorted overallByYear entries
    for p in players_by_name.values():
        ovrs = sorted(p['overallByYear'].items(), key=lambda kv: int(kv[0]))
        p['overallProgression'] = [v for _, v in ovrs]

    players = list(players_by_name.values())
    print(f'\n  → Players merged from per-year rosters: {len(players)}')
    for year in sorted(all_year_data.keys()):
        n = len(all_year_data[year].get('roster') or [])
        print(f'      {year}: {n} on roster')

    # ----- Stamp player.movementByYear from departures -------------------
    for year, yd in all_year_data.items():
        deps = yd.get('departures') or []
        for d in deps:
            p = players_by_name.get(d['name'])
            if not p:
                continue
            mov = {'type': d['movementType']}
            if d['movementType'] == 'transferred_out':
                # Resolve destination abbr → tid when possible
                dest_abbr = d.get('destination')
                mov['toTeamTid'] = tid_for(dest_abbr)
                mov['toTeamAbbr'] = dest_abbr
                mov['reason'] = d.get('reason')
            p['movementByYear'][int(year)] = mov

    # ----- Strip ghost roster entries for players who left -------------
    # The Tracker spreadsheet sometimes carries a player into the next
    # year's Individual Statistics block even after they graduated /
    # got drafted / transferred — the user used those rows as "carry
    # this name into next year so I remember to remove them" notes.
    # On import they show up on the active roster, which is wrong.
    #
    # Rule: if the player has a movement entry of graduated /
    # declared_for_draft / transferred_out in year Y, drop their
    # teamsByYear entries for every year > Y, plus the matching
    # classByYear / overallByYear / statsByYear / devTraitByYear
    # so they don't appear elsewhere either.
    LEAVE_TYPES = ('graduated', 'declared_for_draft', 'transferred_out')
    purged = 0
    for p in players_by_name.values():
        mov = p.get('movementByYear') or {}
        if not mov:
            continue
        # Find the earliest leave year, if any.
        leave_year = None
        for y_raw, m in mov.items():
            try:
                y_int = int(y_raw)
            except (TypeError, ValueError):
                continue
            if m.get('type') in LEAVE_TYPES:
                if leave_year is None or y_int < leave_year:
                    leave_year = y_int
        if leave_year is None:
            continue
        for field in ('teamsByYear', 'classByYear', 'overallByYear',
                      'statsByYear', 'devTraitByYear'):
            store = p.get(field) or {}
            for k in list(store.keys()):
                try:
                    yk = int(k)
                except (TypeError, ValueError):
                    continue
                if yk > leave_year:
                    del store[k]
                    purged += 1
    print(f'  → Purged {purged} ghost roster entries for departed players')

    # ----- Build per-year schedule entries from user's games ------------
    # The Dashboard reads dynasty.teams[userTid].byYear[year].schedule
    # (and falls back to dynasty.schedulesByTeamYear); without a populated
    # schedule the pre-season checklist shows "0/12 games" even when the
    # underlying game records exist.
    schedules_by_team_year = {}
    user_byyear_schedule = {}  # year -> [entries]
    for g in games:
        if g.get('gameType') != 'regular':
            continue
        if g.get('team1') != user_team_abbr and g.get('team2') != user_team_abbr:
            continue
        is_user_team1 = g['team1'] == user_team_abbr
        opp_abbr = g['team2'] if is_user_team1 else g['team1']
        opp_tid = tid_for(opp_abbr)
        location = g.get('location') or 'neutral'
        if location not in ('home', 'away', 'neutral'):
            location = 'neutral'
        entry = {
            'week': g['week'],
            'opponent': opp_abbr,
            'opponentTid': opp_tid,
            'isBye': False,
            'location': location,
            'gameId': g['id'],
            'userTeam': user_team_abbr,
            'userTeamTid': user_team_tid,
        }
        user_byyear_schedule.setdefault(g['year'], []).append(entry)

    for year, sched in user_byyear_schedule.items():
        sched.sort(key=lambda e: e['week'])
        if user_team_tid in dynasty_teams:
            by_year = dynasty_teams[user_team_tid].setdefault('byYear', {})
            year_block = by_year.setdefault(str(year), {})
            year_block['schedule'] = sched
            year_block.setdefault('preseasonSetup', {})['scheduleEntered'] = True
        # Legacy team-centric storage (dual-keyed by abbr + tid)
        schedules_by_team_year.setdefault(user_team_abbr, {})[str(year)] = sched
        schedules_by_team_year.setdefault(str(user_team_tid), {})[str(year)] = sched

    # ----- Stamp byYear[year].record on user team from team_records -----
    if user_team_tid in dynasty_teams:
        by_year = dynasty_teams[user_team_tid].setdefault('byYear', {})
        for year, rec in team_records.items():
            year_block = by_year.setdefault(str(year), {})
            year_block['record'] = {
                'wins': rec['wins'],
                'losses': rec['losses'],
            }

    # ----- Convert games to fully tid-based format -----------------------
    # Drop the legacy abbr fields so the in-app `migrateToFullTidSystem`
    # skip-condition triggers (`team1Tid && team2Tid && !userTeam &&
    # !opponent`) and our pre-resolved tids stay untouched.
    #
    # Also align field names with the app's runtime shape (matching the
    # reference UK_2034 dynasty):
    #   bowl game        : `bowlName` (NOT `bowlGameName`),
    #                      `bowlWeek` = 'week1'/'week2'/...,
    #                      `week` = 'Bowl' (string)
    #   conf championship: `conference` field, `week` = 'CCG' (string)
    #   regular conf game: `isConferenceGame: true` so confWins/confLosses
    #                      get computed
    def conf_for_year(abbr, year):
        """Look up the team's conference for a given year. User team uses
        team_records[year]; everyone else falls back to standings."""
        if abbr == user_team_abbr and year in team_records:
            return team_records[year].get('conference')
        standings = (all_year_data.get(year) or {}).get('standings') or {}
        for conf, teams in standings.items():
            for entry in teams:
                if entry['team'] == abbr:
                    return conf
        return None

    # Build a quick "is user in CFP this year?" lookup. cfp_seeds_by_year
    # isn't populated until the by-year structures pass below, so we
    # walk all_year_data directly here.
    user_in_cfp_year = {}
    for yr_int, yd in all_year_data.items():
        seeds = yd.get('cfpSeeds') or []
        for e in seeds:
            if (e.get('team') or '').upper() == user_team_abbr:
                user_in_cfp_year[int(yr_int)] = True
                break

    # Build a (year, week) → POTW lookup so we can stamp
    # `conferencePOW` / `nationalPOW` on each game record. The Player
    # page sums these per-game stamps to compute career POW totals
    # (no separate by-year structure exists). "Conf Champ" goes on
    # the conference championship game.
    potw_lookup = {}  # {(year, week_int_or_'CCG'): {'conf': name, 'natl': name}}
    for yr_int, yd in all_year_data.items():
        block = yd.get('potw') or {}
        for kind in ('conference', 'national'):
            for entry in block.get(kind, []):
                wk = entry.get('week')
                # Normalize week: numeric stays numeric; "Conf Champ"
                # → 'CCG' (matches the cleaned game's week field for CCs).
                wk_norm = None
                if isinstance(wk, (int, float)):
                    wk_norm = int(wk)
                elif isinstance(wk, str):
                    s_ = wk.strip().lower()
                    if 'champ' in s_:
                        wk_norm = 'CCG'
                    else:
                        try:
                            wk_norm = int(float(s_))
                        except (TypeError, ValueError):
                            wk_norm = None
                if wk_norm is None:
                    continue
                key = (int(yr_int), wk_norm)
                slot = potw_lookup.setdefault(key, {})
                slot['conf' if kind == 'conference' else 'natl'] = entry.get('player')

    tid_games = []
    for g in games:
        t1 = tid_for(g.get('team1'))
        t2 = tid_for(g.get('team2'))
        winner_abbr = g.get('winner')
        winner_tid = tid_for(winner_abbr) if winner_abbr else None
        if g.get('isHome'):
            home_tid = user_team_tid
        elif g.get('isAway'):
            home_tid = t2
        else:
            home_tid = None

        gtype = g['gameType']
        bowl_name_full = g.get('bowlGameName') or ''
        year_int = g['year']

        # If this is a bowl game whose bowl name is one of the
        # CFP-host bowls AND the user was in the CFP that year,
        # promote it from a regular bowl to the matching CFP round.
        # Reassigns gameType and the game id (`{slotId}-{year}`) so
        # CFP Bracket page links resolve.
        if gtype == 'bowl' and bowl_name_full in BOWL_TO_CFP_SLOT and user_in_cfp_year.get(year_int):
            slot_id, cfp_gtype = BOWL_TO_CFP_SLOT[bowl_name_full]
            gtype = cfp_gtype
            game_id = f'{slot_id}-{year_int}'
            is_playoff = True
        elif gtype == 'cfp_championship':
            game_id = f'cfpnc-{year_int}'
            is_playoff = True
        elif gtype == 'cfp_first_round':
            game_id = g['id']  # we don't know which fr1/2/3/4 — leave imported id
            is_playoff = True
        else:
            game_id = g['id']
            is_playoff = g.get('isPlayoff', False)

        clean = {
            'id': game_id,
            'year': year_int,
            'gameType': gtype,
            'team1Tid': t1,
            'team2Tid': t2,
            'team1Score': g.get('team1Score', 0),
            'team2Score': g.get('team2Score', 0),
            'team1Rank': g.get('team1Rank'),
            'team2Rank': g.get('team2Rank'),
            'winnerTid': winner_tid,
            'homeTeamTid': home_tid,
            'isPlayed': g.get('isPlayed', False),
            'isBowlGame': g.get('isBowlGame', False) or gtype in ('cfp_quarterfinal', 'cfp_semifinal', 'cfp_championship'),
            'isConferenceChampionship': g.get('isConferenceChampionship', False),
            'isPlayoff': is_playoff,
        }
        # CFP slot id stamp for bracket lookups
        if gtype in ('cfp_quarterfinal', 'cfp_semifinal', 'cfp_championship'):
            clean['cfpSlot'] = game_id.split('-')[0]

        if gtype in ('bowl', 'cfp_quarterfinal', 'cfp_semifinal', 'cfp_championship'):
            wk_idx = max(1, int(g['week']) - 14) if g.get('week') else 1
            clean['week'] = 'Bowl'
            clean['bowlWeek'] = f'week{wk_idx}'
            clean['bowlName'] = bowl_name_full
        elif gtype == 'cfp_first_round':
            wk_idx = max(1, int(g['week']) - 14) if g.get('week') else 1
            clean['week'] = 'Bowl'
            clean['bowlWeek'] = f'week{wk_idx}'
        elif gtype == 'conference_championship':
            clean['week'] = 'CCG'
            user_conf = (
                team_records.get(g['year'], {}).get('conference')
                or conf_for_year(user_team_abbr, g['year'])
            )
            if user_conf:
                clean['conference'] = user_conf
        else:
            # Regular season — keep numeric week and tag conference games
            # so confWins / confLosses get tallied.
            clean['week'] = g['week']
            user_conf = conf_for_year(user_team_abbr, g['year'])
            opp_abbr = g.get('team2') if g.get('team1') == user_team_abbr else g.get('team1')
            opp_conf = conf_for_year(opp_abbr, g['year'])
            if user_conf and opp_conf and user_conf == opp_conf:
                clean['isConferenceGame'] = True

        # Stamp POTW honorees on the matching game record. The Player
        # page sums conferencePOW / nationalPOW string-equality matches
        # across games — there's no by-year POW table.
        potw_key = (year_int, clean.get('week'))
        potw = potw_lookup.get(potw_key)
        if potw:
            if potw.get('conf'):
                clean['conferencePOW'] = potw['conf']
            if potw.get('natl'):
                clean['nationalPOW'] = potw['natl']

        tid_games.append(clean)
    games = tid_games

    # ----- Build by-year app data structures -----------------------------
    final_polls_by_year = {}
    cfp_seeds_by_year = {}
    conf_standings_by_year = {}
    conf_champs_by_year = {}
    natl_champ_by_year = {}
    awards_by_year = {}
    all_americans_by_year = {}
    players_of_week_by_year = {}
    players_leaving_by_year = {}
    players_leaving_by_team_year = {}
    recruits_by_team_year = {}
    recruiting_commitments_by_team_year = {}  # primary path the app reads
    legacy_recruits = []

    team_records_by_team_year = {}   # abbr -> { year -> {wins, losses} }
    conf_by_team_year = {}           # abbr -> { year -> conference }

    for year, yd in all_year_data.items():
        ystr = str(year)

        if yd.get('finalPolls'):
            # Strip the placeholder tid:null fields and resolve real
            # tids — matches the reference dynasty's `{rank, team, tid}`
            # poll entry shape.
            polls = yd['finalPolls']
            polls_clean = {
                k: [
                    {'rank': e['rank'], 'team': e['team'], 'tid': tid_for(e['team'])}
                    for e in (polls.get(k) or [])
                ]
                for k in ('media', 'coaches')
            }
            final_polls_by_year[ystr] = polls_clean
        if yd.get('cfpSeeds'):
            # Reference shape: {seed, team, tid} only — no conference/autoBid
            cfp_seeds_by_year[ystr] = [
                {'seed': e['seed'], 'team': e['team'], 'tid': tid_for(e['team'])}
                for e in yd['cfpSeeds']
            ]
        if yd.get('standings'):
            conf_standings_by_year[ystr] = yd['standings']
            # Stamp team records + conference for every team in standings
            for conf, teams in yd['standings'].items():
                for entry in teams:
                    abbr = entry['team']
                    team_records_by_team_year.setdefault(abbr, {})[ystr] = {
                        'wins': entry['wins'],
                        'losses': entry['losses'],
                        'pointsFor': entry.get('pointsFor'),
                        'pointsAgainst': entry.get('pointsAgainst'),
                        'pointDiff': entry.get('pointDiff'),
                        'lastUpdated': '',
                    }
                    conf_by_team_year.setdefault(abbr, {})[ystr] = conf
        if yd.get('confChamps'):
            # Reference shape: {conference, team1, team2, team1Score,
            # team2Score, winner}. Our parser gives {conference,
            # champion, opponent, coach}. We don't have CC scores in
            # the standings table (only the user's CC has a score, in
            # the schedule grid). Leave team1Score/team2Score = 0 — the
            # app reads scores from games[] anyway via the CC merge.
            cc_records = []
            for cc in yd['confChamps']:
                champ = cc.get('champion')
                opp = cc.get('opponent')
                rec = {
                    'conference': cc.get('conference'),
                    'team1': champ,
                    'team2': opp,
                    'team1Score': 0,
                    'team2Score': 0,
                    'winner': champ,
                }
                if cc.get('coach'):
                    rec['coach'] = cc['coach']
                cc_records.append(rec)
            conf_champs_by_year[ystr] = cc_records
        if yd.get('natlChamp'):
            natl_champ_by_year[ystr] = yd['natlChamp']
        if yd.get('awards'):
            awards_by_year[ystr] = yd['awards']

        # Combine AA + AC into allAmericansByYear (the schema the app uses)
        aa_entry = {}
        if yd.get('allAmericans'):
            aa_entry['allAmericans'] = yd['allAmericans']
        if yd.get('allConference'):
            aa_entry['allConference'] = yd['allConference']
            # Group AC by school's conference for the optional
            # allConferenceByConference index. We use the user's
            # conference standings to map school → conference.
            standings = yd.get('standings') or {}
            school_to_conf = {}
            for conf, teams in standings.items():
                for t in teams:
                    school_to_conf[t['team']] = conf
            grouped = {}
            for entry in yd['allConference']:
                school = entry.get('school')
                conf = school_to_conf.get(school) if school else None
                if conf:
                    grouped.setdefault(conf, []).append(entry)
            if grouped:
                aa_entry['allConferenceByConference'] = grouped
        if aa_entry:
            all_americans_by_year[ystr] = aa_entry

        if yd.get('potw'):
            players_of_week_by_year[ystr] = yd['potw']

        # Player departures → both shapes
        deps = yd.get('departures') or []
        if deps:
            simple = [
                {
                    'playerName': d['name'],
                    'pid': (players_by_name.get(d['name']) or {}).get('pid'),
                    'reason': d.get('reason') or 'Unknown',
                    'destination': d.get('destination'),
                    'overall': d.get('overall'),
                    'position': d.get('position'),
                    'class': d.get('class'),
                    'draftRound': d.get('draftRound'),
                }
                for d in deps
            ]
            players_leaving_by_year[ystr] = simple
            players_leaving_by_team_year.setdefault(user_team_abbr, {})[ystr] = simple

        # Recruits — the Recruiting page reads
        # `recruitingCommitmentsByTeamYear[teamKey][year][weekKey] = [recruits]`.
        # Use 'signing_5' as the bucket label (matches how completed
        # classes are stored in the reference dynasty). Each recruit
        # entry uses the canonical recruit shape (no firstName/lastName
        # split, just `name`).
        rec = yd.get('recruits') or []
        if rec:
            recruits_by_team_year.setdefault(user_team_abbr, {})[ystr] = rec
            commitments_year = {
                'signing_5': [
                    {
                        'name': r.get('name', ''),
                        'class': r.get('class') or 'HS',
                        'position': r.get('position') or '',
                        'archetype': r.get('archetype') or '',
                        'stars': r.get('stars'),
                        'nationalRank': r.get('nationalRank'),
                        'stateRank': r.get('stateRank'),
                        'positionRank': r.get('positionRank'),
                        'height': r.get('height') or '',
                        'weight': r.get('weight'),
                        'hometown': r.get('hometown') or '',
                        'state': r.get('state') or '',
                        'gemBust': r.get('gemBust') or '',
                        'devTrait': r.get('devTrait') or 'Normal',
                        'previousTeam': r.get('previousTeam') or '',
                        'isPortal': r.get('isPortal', False),
                    }
                    for r in rec
                ]
            }
            recruiting_commitments_by_team_year.setdefault(user_team_abbr, {})[ystr] = commitments_year
            for r in rec:
                legacy_recruits.append({**r, 'team': user_team_abbr, 'recruitYear': year})

    # ----- Coach career ---------------------------------------------------
    coach_team_by_year = {}
    coach_career = []
    for year, rec in team_records.items():
        # coachTeamByYear shape (matches reference):
        # { tid, team (abbr), teamName, position }
        coach_team_by_year[str(year)] = {
            'tid': user_team_tid,
            'team': user_team_abbr,
            'teamName': user_team_full_name,
            'position': rec['role'],
        }
        # coachCareer shape (matches reference): one entry per year, just
        # { tid, year, position }. The CoachCareer page derives wins,
        # losses, and conference from games[] / conferenceByTeamYear.
        coach_career.append({
            'tid': user_team_tid,
            'year': year,
            'position': rec['role'],
        })
        # User team record → if not already in team_records_by_team_year
        # (which comes from standings), drop in conferred values from Team sheet
        if user_team_abbr not in team_records_by_team_year or str(year) not in team_records_by_team_year[user_team_abbr]:
            team_records_by_team_year.setdefault(user_team_abbr, {})[str(year)] = {
                'wins': rec['wins'],
                'losses': rec['losses'],
                'lastUpdated': '',
            }
        if rec['conference']:
            conf_by_team_year.setdefault(user_team_abbr, {}).setdefault(str(year), rec['conference'])

    # ----- Link recruits to players (shared pid) ------------------------
    # The user expects each recruit on the Recruiting page to share a
    # pid with the same-named player on the active roster, just like
    # if they had been entered through the site normally. We do this
    # AFTER all year-rosters are merged (so every recurring name is
    # already in players_by_name), and AFTER coach career build (so
    # we don't disturb earlier ordering).
    #
    # Strategy: for each recruit, look up by exact name in
    # players_by_name. If found → stamp `pid` on the recruit entry
    # AND copy recruit-specific fields (archetype, height, weight,
    # hometown, state, gemBust, isPortal, previousTeam, recruitYear,
    # stars, nationalRank) onto the player record. If not found
    # (recruit who never made the roster), spin up a new player
    # record with isHonorOnly=False, isRecruit=True so they're
    # tracked but not surfaced as active roster.
    matched, created = 0, 0
    for team_abbr, by_year in recruiting_commitments_by_team_year.items():
        for year_str, year_buckets in by_year.items():
            try:
                recruit_year_int = int(year_str)
            except (TypeError, ValueError):
                recruit_year_int = None
            for week_key, recruit_list in year_buckets.items():
                for r in recruit_list:
                    name = (r.get('name') or '').strip()
                    if not name:
                        continue
                    if name in players_by_name:
                        p = players_by_name[name]
                        r['pid'] = p['pid']
                        # Copy recruit metadata onto player (don't
                        # overwrite already-populated fields — roster
                        # data wins for things like position).
                        if not p.get('archetype'):
                            p['archetype'] = r.get('archetype', '')
                        if not p.get('height'):
                            p['height'] = r.get('height', '')
                        if not p.get('weight'):
                            p['weight'] = r.get('weight')
                        if not p.get('hometown'):
                            p['hometown'] = r.get('hometown', '')
                        if not p.get('state'):
                            p['state'] = r.get('state', '')
                        if not p.get('gemBust'):
                            p['gemBust'] = r.get('gemBust', '')
                        # Stars / nationalRank from Individual sheet
                        # might be missing; recruit row has them.
                        if p.get('stars') is None:
                            p['stars'] = r.get('stars')
                        if p.get('nationalRank') is None:
                            p['nationalRank'] = r.get('nationalRank')
                        if not p.get('previousTeam'):
                            p['previousTeam'] = r.get('previousTeam') or ''
                        p['isRecruit'] = True
                        p['isPortal'] = bool(r.get('isPortal'))
                        if recruit_year_int is not None:
                            p['recruitYear'] = recruit_year_int
                            # If the player has a yearStarted later than
                            # recruitYear (e.g. roster data wasn't filled
                            # in for an early year), keep the older one.
                            ys = p.get('yearStarted')
                            if not ys or recruit_year_int < ys:
                                p['yearStarted'] = recruit_year_int
                        matched += 1
                    else:
                        # Recruit didn't show up on any year's Individual
                        # Statistics roster. Still create a player record
                        # so they aren't orphaned — that mirrors what
                        # happens in normal site flow (recruits become
                        # roster players on year-rollover) AND propagate
                        # them forward through subsequent seasons until
                        # graduation, since users should see continuity
                        # ('28 recruit shows up on '29 roster as a So,
                        # etc.) without having to re-enter them.
                        first, last = split_name(name)
                        pid = next_pid
                        next_pid += 1
                        new_p = {
                            'pid': pid,
                            'id': f'imported-player-{pid}',
                            'name': name,
                            'firstName': first,
                            'lastName': last,
                            'position': r.get('position') or '',
                            'team': user_team_tid,
                            'school': team_abbr,
                            'year': r.get('class') or 'Fr',
                            'stars': r.get('stars'),
                            'nationalRank': r.get('nationalRank'),
                            'devTrait': r.get('devTrait') or 'Normal',
                            'devTraitByYear': {},
                            'overall': None,
                            'overallByYear': {},
                            'overallProgression': [],
                            'teamsByYear': {},
                            'classByYear': {},
                            'statsByYear': {},
                            'movementByYear': {},
                            'movements': [],
                            'yearStarted': recruit_year_int or start_year,
                            'jerseyNumber': '',
                            'archetype': r.get('archetype') or '',
                            'height': r.get('height') or '',
                            'weight': r.get('weight'),
                            'hometown': r.get('hometown') or '',
                            'state': r.get('state') or '',
                            'gemBust': r.get('gemBust') or '',
                            'pictureUrl': '',
                            'isHonorOnly': False,
                            'isPortal': bool(r.get('isPortal')),
                            'isRecruit': True,
                            'recruitYear': recruit_year_int,
                            'previousTeam': r.get('previousTeam') or '',
                        }
                        # Propagate roster membership year-over-year
                        # using the class-progression table.
                        if recruit_year_int is not None:
                            initial_class = CLASS_PROGRESSION.get(
                                r.get('class') or 'HS', r.get('class') or 'Fr'
                            )
                            cur_class = initial_class
                            cur_year = recruit_year_int
                            while cur_year <= last_active_year and cur_class:
                                new_p['teamsByYear'][str(cur_year)] = user_team_tid
                                new_p['classByYear'][str(cur_year)] = cur_class
                                cur_class = NEXT_CLASS.get(cur_class)
                                cur_year += 1
                        players_by_name[name] = new_p
                        r['pid'] = pid
                        created += 1
    # Refresh the players list since we may have appended new entries.
    players = list(players_by_name.values())
    print(f'  → Recruit↔player link: {matched} matched to roster, {created} new players created')

    # Mirror pids onto the legacy `recruits` list too (same rules).
    name_to_pid = {n: p['pid'] for n, p in players_by_name.items()}
    for r in legacy_recruits:
        nm = (r.get('name') or '').strip()
        if nm and nm in name_to_pid:
            r['pid'] = name_to_pid[nm]

    # ----- Build dynasty skeleton ----------------------------------------
    # Final-form team name (from TB metadata if user is on a TB,
    # otherwise from Team sheet text). dynasty.teams[userTid].name is
    # the canonical display name; teamName here mirrors it.
    user_team_display_name = (
        dynasty_teams.get(user_team_tid, {}).get('name') or user_team_full_name
    )

    # ----- Build customConferences map from user's standings ------------
    # The user did extensive realignment. Reference dynasty stores this
    # as customConferences (current snapshot, conf → [abbr,...]) and
    # customConferencesByYear (year-stamped historical view).
    custom_conferences_by_year = {}  # year-stamped snapshots
    most_complete_year = None
    most_complete_count = -1
    for year in sorted(all_year_data.keys()):
        standings = all_year_data[year].get('standings') or {}
        if not standings:
            continue
        year_snapshot = {
            conf: [e['team'] for e in entries if e.get('team')]
            for conf, entries in standings.items()
        }
        # Year-stamped record always uses the year's own snapshot.
        custom_conferences_by_year[str(year)] = year_snapshot
        # Track which year has the most conferences populated — that's
        # the user's intended realignment snapshot.
        team_count = sum(len(v) for v in year_snapshot.values())
        if team_count > most_complete_count:
            most_complete_count = team_count
            most_complete_year = year

    # Live snapshot = most-complete year's mapping. If no year has any
    # standings, leave empty (the static FBS map kicks in as fallback).
    custom_conferences = (
        custom_conferences_by_year.get(str(most_complete_year), {})
        if most_complete_year is not None else {}
    )

    # ----- Multiplayer-of-N schema -------------------------------------
    # We DON'T pre-seed editors / memberTeams / memberLabels /
    # memberTeamHistory here — those need to be keyed by the importer's
    # real Firebase auth UID, which we obviously don't know at
    # migration time. processImportData() in DynastyContext.jsx seeds
    # them post-import.
    #
    # We do prepare a year-by-tid map, though, so the import seeder
    # can build memberTeamHistory directly without re-deriving from
    # coachTeamByYear:
    member_team_history_by_year = {
        str(year): [user_team_tid] for year in sorted(team_records.keys())
    }

    dynasty = {
        'name': f'{user_team_abbr} Dynasty (imported)',
        'dynastyName': f'{user_team_abbr} Dynasty (imported)',
        'teamName': user_team_display_name,
        'currentTid': user_team_tid,
        'coachName': coach_name,
        'coachPosition': coach_position,
        'startYear': start_year,
        # Land at 2029 Week 12 so the user picks up mid-season on the
        # last roster they had filled in.
        'currentYear': last_active_year,
        'currentWeek': 12,
        'currentPhase': 'regular_season',
        'conference': team_records.get(last_active_year, {}).get('conference') or '',
        'storageType': 'local',

        # TB info baked directly into dynasty.teams[tid] — the site reads
        # the slot and renders whatever's there. No separate TB concept.
        'teams': {str(t): v for t, v in dynasty_teams.items()},
        # Marker so the in-app `customTeams → dynasty.teams` collapse
        # migration is a no-op.
        '_tidMigrated': True,
        '_tidFullyMigrated': True,

        # Year-keyed structured data
        'games': games,
        'players': players,
        'recruits': legacy_recruits,
        'recruitsByTeamYear': recruits_by_team_year,
        'recruitingCommitmentsByTeamYear': recruiting_commitments_by_team_year,
        'finalPollsByYear': final_polls_by_year,
        'cfpSeedsByYear': cfp_seeds_by_year,
        'conferenceStandingsByYear': conf_standings_by_year,
        'conferenceChampionshipsByYear': conf_champs_by_year,
        'nationalChampionByYear': natl_champ_by_year,
        'awardsByYear': awards_by_year,
        'allAmericansByYear': all_americans_by_year,
        'playersOfTheWeekByYear': players_of_week_by_year,
        'playersLeavingByYear': players_leaving_by_year,
        'playersLeavingByTeamYear': players_leaving_by_team_year,
        'schedulesByTeamYear': schedules_by_team_year,

        'schedule': user_byyear_schedule.get(last_active_year, []),
        'rankings': [],
        'nextPID': next_pid,

        'preseasonSetup': {
            'scheduleEntered': True,
            'rosterEntered': True,
            'teamRatingsEntered': False,
            'coachingStaffEntered': False,
            'conferencesEntered': False,
        },
        'teamRatings': {'overall': None, 'offense': None, 'defense': None},
        'coachingStaff': {
            'hcName': coach_name if coach_position == 'HC' else None,
            'ocName': coach_name if coach_position == 'OC' else None,
            'dcName': coach_name if coach_position == 'DC' else None,
        },
        'coachCareer': coach_career,
        'coachTeamByYear': coach_team_by_year,
        'teamRecordsByTeamYear': team_records_by_team_year,
        'conferenceByTeamYear': conf_by_team_year,

        # Conference realignment
        'customConferences': custom_conferences,
        'customConferencesByYear': custom_conferences_by_year,

        # Hint for the import seeder. The actual editors / memberTeams /
        # memberLabels / memberTeamHistory entries are written by
        # processImportData() once the importer's real UID is known.
        '_importMemberSeed': {
            'tid': user_team_tid,
            'coachName': coach_name,
            'teamHistoryByYear': member_team_history_by_year,
        },

        # Migration markers — set everything so the in-app migration
        # passes are no-ops on first load.
        '_schemaVersion': 2,
        '_tidDataMigrationPending': False,
        '_userTeamSystemMigrated': True,
        '_coachTeamByYearMigrated': True,
        '_movementsMigrated': True,
        '_legacyDepartureFieldsMigrated': True,
        '_gamesMigrated': True,
        '_statsMigrated': True,
        '_subcollectionsMigrated': False,  # set True only after a real Firestore migration
        '_normalizedAt': '',
        '_gameCount': len(games),
        '_playerCount': len(players),

        # Misc state
        'isPublic': False,
        'favorite': False,
        'seasons': [],
        'coachingHistory': [],
        'isFirstYearOnCurrentTeam': False,
        'newJobData': None,
        'previousJobData': None,
        'pendingCoordinatorHires': None,
    }

    # ----- Output --------------------------------------------------------
    print(f'\nWriting {output_path}…')
    output_path.write_text(json.dumps(dynasty, indent=2, default=str))
    size_kb = output_path.stat().st_size / 1024
    print(f'  → {size_kb:,.1f} KB')
    print('\nSummary:')
    print(f'  Games: {len(games)}')
    print(f'  Players: {len(players)}')
    print(f'  Slots overridden in dynasty.teams: {len(dynasty_teams)} '
          f'(user team = {user_team_abbr} → tid {user_team_tid})')
    print(f'  Years with standings: {len(conf_standings_by_year)}')
    print(f'  Years with awards: {len(awards_by_year)}')
    print(f'  Years with All-Americans/Conference: {len(all_americans_by_year)}')
    print(f'  Years with departures: {len(players_leaving_by_year)}')
    print(f'  Years with recruiting class: '
          f'{len(recruits_by_team_year.get(user_team_abbr, {}))}')
    print('\nDone. Import this file via the homepage "Import File" button.')
    print('  • Coach name is "[Your Name]" — set it via Account')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('xlsx', type=Path, help='Path to the Tracker .xlsx')
    parser.add_argument('output', nargs='?', type=Path, default=None,
                        help='Output JSON path (default: <xlsx-name>-migrated.json)')
    args = parser.parse_args()

    if not args.xlsx.exists():
        sys.exit(f'File not found: {args.xlsx}')
    output = args.output or args.xlsx.with_suffix('').with_name(args.xlsx.stem + '-migrated.json')
    migrate(args.xlsx, output)
